# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio

import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file, jsonify

# guardar cliente por primera y unica vez
def procesar_form_cliente(dataForm, foto_perfil):
    uno=dataForm['id']
    dos=dataForm['nombre_empleado']
    tres=dataForm['apellido_empleado']
    cuatro=dataForm['sexo_empleado']
    cinco=dataForm['telefono_empleado']
    seis=dataForm['ubicacion']
    siete=dataForm['email_empleado']
    ocho=procesar_imagen_perfil(foto_perfil)
    nueve='1'
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "INSERT INTO tbl_empleados( id, nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, ubicacion, email_empleado, foto_empleado, estado)  VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (uno, dos, tres, cuatro, cinco, seis, siete, ocho, nueve)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert
    except Exception as e:
        return f'Se produjo un error en procesar_form_cliente: {str(e)}'
    
# guardar documento de cada proceso
def procesar_form_carga(dataForm, document, tipo, tproceso):
    result_carga_document = procesar_carga_document(document, tipo, tproceso)
    result_filename = procesar_nombre_filename(document)
    estado = '1'
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "INSERT INTO tbl_carga_document( id_cliente, nombre, apellido, id_proceso, id_etapa, document, filename, estado)  VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (dataForm['id_cliente'], dataForm['nombre_empleado'], dataForm['apellido_cliente'], dataForm['id_proceso'], dataForm['id_etapa_etapa'], result_carga_document, result_filename, estado)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert
    except Exception as e:
        return f'Se produjo un error al cargar un documento de un proceso: {str(e)}'

# guardar nuevo procesos de un cliente
def procesar_form_proceso(dataForm):
    id=dataForm['id_cliente']
    idti=dataForm['proce']
    proceso_numero=dataForm['proceso_numero']
    demandado=dataForm['demandado']
    depto=dataForm['depto']
    id_juzgado=dataForm['id_juzgado']
    estado = '1'
    # asignandole valores a los input deshabilitados
    if idti == '1':
        #ar=dataForm['areas']
        id_area_proceso='24'
    else:
        id_area_proceso=dataForm['areas']
    if idti==3:
        id_area_proceso='24'
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "INSERT INTO tbl_procesos( id_cliente, id_tipo_proceso, id_area_proceso, proceso_numero, id_depto, id_juzgado, demandado, estado )  VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (id, idti, id_area_proceso, proceso_numero, depto, id_juzgado, demandado, estado)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado = cursor.rowcount
                return resultado
    except Exception as e:
        return f'Se produjo un error al guardar nuevo proceso de un cliente: {str(e)}'

# guardar en tabla proceso delitos al momento de guardar un proceso por primera vez
def procesar_form_proceso_delito(dataForm, resulProceso):
    proce=dataForm['proce']
    if(proce=='1'):
       id_delito='0'
    else:
        id_delito='287'
    id=dataForm['id_cliente']
    proceso_numero=resulProceso
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "INSERT INTO tbl_proceso_delito( id_delito, id_cliente, id_proceso_numero)  VALUES (%s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (id_delito, id, proceso_numero)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                res= cursor.rowcount
                return res
    except Exception as e:
        return f'Se produjo un error al guardar en tabla proceso delito: {str(e)}'

# consulta para traer el numero de proceso que asigno el sistema esto para guardar por primera vez en delitos de un proceso
def resul_Proceso(dataForm):
    pro_numero=dataForm['proceso_numero']
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = ("""
                            SELECT 
                                id_proceso
                            FROM 
                                tbl_procesos 
                          	WHERE proceso_numero =  %s
                            """)
                cursor.execute(sql, (pro_numero,))
                re = cursor.fetchone()
                for dato, valor in re.items():
                    r2=valor
                return r2
    except Exception as e:
        return f'Se produjo un error al consultar el numero de proceso: {str(e)}'


# consulta para traer el numero de proceso que asigno el sistema esto para guardar por segunda o mas veces un delito
def resul_Proceso2(p_numero):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = ("""
                            SELECT 
                                id_proceso
                            FROM 
                                tbl_procesos 
                          	WHERE proceso_numero =  %s
                            """)
                cursor.execute(sql, (p_numero,))
                re = cursor.fetchone()
                for dato, valor in re.items():
                    r2=valor
                return r2
    except Exception as e:
        return f'Se produjo un error al consultar el numero de proceso para listarlo en delitos: {str(e)}'


# guardar delitos de un proceso y de un cliente
def procesar_form_delito(dataForm, r2):
    id=dataForm['id_cliente']
    id_delito=dataForm['id_delito']
    agra=dataForm['agra']
    proceso_numero=r2
    estado = '1'
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "INSERT INTO tbl_delitos_agravantes( id_cliente, id_delito, id_agravante, id_proceso_delagra, estado )  VALUES (%s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (id, id_delito, agra, proceso_numero, estado)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado = cursor.rowcount
                return resultado
    except Exception as e:
        return f'Se produjo un error al guardar nuevo delito de un cliente: {str(e)}'

# guardar en tabla proceso_delitos si el tipo de proceso es penal
def procesar_form_delito_update(dataForm, r2):
    id_delito=dataForm['id_delito']
    id=dataForm['id_cliente']
    proceso_numero=r2
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "UPDATE tbl_proceso_delito SET id_delito=%s, id_cliente=%s WHERE id_proceso_numero = %s"
                # Creando una tupla con los valores del INSERT
                valores = (id_delito, id, proceso_numero)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado2 = cursor.rowcount
                return resultado2
    except Exception as e:
        return f'Se produjo un error al actualizar delito: {str(e)}'

# guardar etapas de los procesos
def procesar_form_etapa(dataForm):
    estado = '1'
    id_eta_eta=dataForm['id_etapa_etapa']
    id_cliente=dataForm['id_cliente']
    anot_etapa=dataForm['anotacion_etapa']
    id_pro=dataForm['id_proceso']
    faudi=dataForm['fau']
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = "INSERT INTO tbl_new_etapa( id_etapa_etapa, id_cliente, anotacion_etapa, estado_new_etapa, id_proceso_etapa, fecha_audiencia)  VALUES (%s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (id_eta_eta, id_cliente, anot_etapa, estado, id_pro, faudi)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado = cursor.rowcount
                return resultado, id_pro
    except Exception as e:
        return f'Se produjo un error en procesar nueva etapa de un proceso: {str(e)}'


#guardar el documento de identidad de un cliente
def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteresbuscarClienteUnico
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/id_clientes/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []

#carga de documento de un proceso
def procesar_carga_document(docu, filename, tproceso):
    try:
        # Nombre original del archivo
        filename = secure_filename(docu.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteresbuscarClienteUnico
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        # basepath = os.path.abspath(os.path.dirname(__file__))
        # upload_dir = os.path.join(basepath, f'../static/id_clientes/')
        if tproceso == '1':
            basepath = os.path.abspath(os.path.dirname(__file__))
            upload_dir = os.path.join(basepath, f'../static/doc_penal/')
        elif tproceso == '2':
            basepath = os.path.abspath(os.path.dirname(__file__))
            upload_dir = os.path.join(basepath, f'../static/doc_administrativo/')
        elif tproceso == '3':
            basepath = os.path.abspath(os.path.dirname(__file__))
            upload_dir = os.path.join(basepath, f'../static/doc_laboral/')
        elif tproceso == '4':
            basepath = os.path.abspath(os.path.dirname(__file__))
            upload_dir = os.path.join(basepath, f'../static/doc_civil/')
        elif tproceso == '5':
            basepath = os.path.abspath(os.path.dirname(__file__))
            upload_dir = os.path.join(basepath, f'../static/doc_familia/')    
        
        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        docu.save(upload_path)
        return nombreFile
    except Exception as e:
        print("Error al guardar el documento en la carpeta:", e)
        return []

#obtener nombre original del archivo
def procesar_nombre_filename(docu):
    try:
        # Nombre original del archivo
        filename = secure_filename(docu.filename)
        return filename
    except Exception as e:
        print("Error al guardar el documento en la carpeta:", e)
        return []

# Lista de clientes
def sql_lista_clientesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.id,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                clientesBD = cursor.fetchall()
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_clientesBD: {e}")
        return None

# Lista de procesos por cliente
def sql_lista_procesosBD(idCliente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                            SELECT DISTINCT
                                *
                            FROM 
                                tbl_empleados AS e, tbl_procesos AS p, tbl_tipo_proceso AS t
                            WHERE
                                e.id_empleado = %s AND e.id=p.id_cliente AND p.id_tipo_proceso=t.id_tipo_proceso AND p.estado=1
                            ORDER BY id_proceso DESC
                    """)
                cursor.execute(querySQL, (idCliente,))
                procesosBD = cursor.fetchall()
        return procesosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_procesosBD: {e}")
        return None

# Lista de etapas por proceso
def sql_lista_etapasBD(idEtapa):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                      e.id_new_etapa, e.id_cliente, e.id_new_etapa,
                      e.id_proceso_etapa, e.anotacion_etapa, e.estado_new_etapa,
                      p.id_proceso, p.id_tipo_proceso, t.id_tipo_proceso
                    FROM 
                        tbl_new_etapa AS e, tbl_procesos AS p, tbl_tipo_proceso AS t
                    WHERE
                       p.id_proceso = e.id_proceso_etapa AND t.id_tipo_proceso = p.id_tipo_proceso AND
                       		e.estado_new_etapa = 1 AND e.id_proceso_etapa = %s 
                    ORDER BY e.id_new_etapa DESC
                    """)
                cursor.execute(querySQL, (idEtapa,))
                etapasBD = cursor.fetchall()
        return etapasBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_etpasBD: {e}")
        return None

# Detalles del cliente
def sql_detalles_clientesBD(idCliente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.id,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idCliente,))
                clientesBD = cursor.fetchone()
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_clientesBD: {e}")
        return None

# Funcion Clientes Informe (Reporte)
def clientesReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.id,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                clientesBD = cursor.fetchall()
        return clientesBD
    except Exception as e:
        print(
            f"Error en la función clientesReporte: {e}")
        return None

def generarReporteExcel():
    dataClientes = clientesReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active
    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataClientes:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, 
                    fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

#Buscando clientes en tabla de busqueda
def buscarClienteBD(search):
    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.id,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.estado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id LIKE %s AND e.estado = 1
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscar un cliente: {e}")
        return []

#Buscando procesos en tabla de busqueda
def buscarProcesoBD(search, id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                if search :    
                    querySQL = ("""
                                SELECT DISTINCT
                                    c.id, c.nombre_empleado, c.apellido_empleado,
                                    p.id_cliente, p.id_tipo_proceso, p.id_juzgado, p.proceso_numero, p.estado, 												p.id_proceso,
                                    t.proceso_nombre, t.id_tipo_proceso
                                FROM tbl_empleados AS c, tbl_procesos AS p, tbl_tipo_proceso AS t
                                WHERE p.proceso_numero LIKE %s AND p.estado = 1 AND c.id = 												p.id_cliente AND p.id_tipo_proceso = t.id_tipo_proceso
                                       AND p.id_cliente=%s
                            """)
                else:
                    querySQL = ("""
                                SELECT DISTINCT
                                    c.id, c.nombre_empleado, c.apellido_empleado,
                                    p.id_cliente, p.id_tipo_proceso, p.id_juzgado, p.proceso_numero, p.estado, 												p.id_proceso,
                                    t.proceso_nombre, t.id_tipo_proceso
                                FROM tbl_empleados AS c, tbl_procesos AS p, tbl_tipo_proceso AS t
                                WHERE p.proceso_numero LIKE %s AND p.estado = 1 AND c.id = 												p.id_cliente AND p.id_tipo_proceso = t.id_tipo_proceso
                                       AND p.id_cliente=%s
                            """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern, id,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscar un proceso de un cliente: {e}")
        return []


#Buscando etapas en tabla de busqueda
def buscarEtapaBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_new_etapa, e.id_cliente, e.id_new_etapa,
                            e.id_proceso_etapa, e.id_etapa_etapa, e.anotacion_etapa, e.estado_new_etapa,
                            p.id_proceso, p.id_tipo_proceso, t.id_tipo_proceso, t.id_tipo_area, p.id_cliente,
                            i.nombre_etapa, i.id_etapa,
                            a.id_area_tip
                            
                        FROM 
                            tbl_new_etapa AS e, tbl_procesos AS p, tbl_tipo_proceso AS t, tbl_etapas AS i, tbl_tipo_area AS a
                        WHERE
                            p.id_proceso = e.id_proceso_etapa AND t.id_tipo_proceso = p.id_tipo_proceso AND
                       		e.estado_new_etapa = 1 AND t.id_tipo_area=a.id_area_tip AND p.id_cliente=e.id_cliente AND e.id_new_etapa = i.id_etapa AND i.nombre_etapa = %s 
                        ORDER BY e.id_new_etapa DESC
                        
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscar una etapa de un proceso: {e}")
        return []

#buscar cliente unico en tabla de clientes
def buscarClienteUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.id,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.ubicacion,
                            e.email_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1 
                    """)
                mycursor.execute(querySQL, (id,))
                cliente = mycursor.fetchone()
                return cliente
    except Exception as e:
        print(f"Ocurrió un error en def buscarClienteUnico: {e}")
        return []

#buscar proceso unico en tabla de procesos para hacer la edicion de un proceso
def buscarProUnico(ti):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT DISTINCT
                            p.id_proceso, p.id_cliente, p.id_tipo_proceso, p.id_juzgado, p.proceso_numero, p.estado, p.id_area_proceso, p.demandado, 	  p.id_depto,
                            e.id, e.nombre_empleado, e.apellido_empleado,
                            t.id_tipo_proceso, t.proceso_nombre,
                            j.id_juzgado, j.nombre_juzgado,
                            a.id_area_tip, a.nombre_area,
                            m.nombre, m.cod_municipio
                        FROM
                            tbl_procesos AS p, tbl_empleados AS e, tbl_tipo_proceso AS t, tbl_juzgados AS j, tbl_tipo_area  AS a, tbl_municipios AS m
                        WHERE
                            p.id_proceso = %s AND p.id_cliente=e.id AND p.id_tipo_proceso=t.id_tipo_proceso AND p.id_juzgado=j.id_juzgado 
                            AND p.id_area_proceso=a.id_tip_area AND p.id_depto=m.cod_municipio LIMIT 1 
                    """)
                mycursor.execute(querySQL, (ti,))
                pro = mycursor.fetchone()
                return pro
    except Exception as e:
        print(f"Ocurrió un error en def buscar proceso unico para edicion: {e}")
        return []

# seleccionar agravantes de un de proceso para llenar el listado de delitos y agravantes al editar un proceso
def agraVantes(ti):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        	SELECT DISTINCT
                            p.id_proceso, p.proceso_numero, n.id_proceso_delagra, n.id_delito, n.id_agravante, n.id_delagra,
                            d.id_delito, d.delitos_nombre,
                            a.id_agravante, a.descripcion, a.id_delito_agravante, s.id_proceso_numero
                        FROM
                            tbl_procesos AS p, tbl_delitos_agravantes AS n, tbl_delitos AS d, tbl_agravantes AS a, tbl_proceso_delito AS s
                        WHERE
                            n.id_proceso_delagra =%s AND p.id_proceso=n.id_proceso_delagra AND d.id_delito=n.id_delito AND
                            p.id_proceso=s.id_proceso_numero AND n.id_agravante=a.id_agravante AND n.estado = 1
                    """)
                mycursor.execute(querySQL, (ti,))
                agr = mycursor.fetchall()
                return agr
    except Exception as e:
        print(f"Ocurrió un error en def buscar agravantes de delitos para editar : {e}")
        return []

#buscar proceso unico
def buscarProcesoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            *
                        FROM tbl_procesos, tbl_empleados
                        WHERE  id_proceso =%s LIMIT 1 
                    """)
                mycursor.execute(querySQL, (id,))
                proceso = mycursor.fetchone()
                return proceso
    except Exception as e:
        print(f"Ocurrió un error en def buscar un proceso de un z: {e}")
        return []

#buscar proceso unico para el buscador de procesos
def buscarProcesoBus(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            *
                        FROM tbl_procesos, tbl_empleados
                        WHERE id_cliente=id AND id_empleado =%s LIMIT 1 
                    """)
                mycursor.execute(querySQL, (id,))
                pro = mycursor.fetchone()
                if (pro == None):
                    proceso='1' 
                else:
                    proceso=pro  
                return proceso
    except Exception as e:
        print(f"Ocurrió un error en def buscar un proceso de un z: {e}")
        return []

#buscar proceso unico para carga de documentos    
def buscarProcesoUnicoDocu(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            *
                        FROM tbl_procesos, tbl_empleados
                        WHERE id_proceso =%s LIMIT 1 
                    """)
                mycursor.execute(querySQL, (id,))
                proceso = mycursor.fetchone()
                return proceso
    except Exception as e:
        print(f"Ocurrió un error en def buscar un proceso de un z: {e}")
        return []

#cargando combo de tipo de procesos en el formulario nuevo proceso del cliente
def cargaCombotippro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursorr:
                querySQL1 = ("""
                        SELECT 
                           *
                        FROM tbl_tipo_proceso
                        WHERE proceso_estado = 1 
                    """)
                mycursorr.execute(querySQL1)
                tipo = mycursorr.fetchall()
                return tipo
    except Exception as e:
        print(f"Ocurrió un error al hacer la consulta de tipo de proceso: {e}")
        return []

# cargando combo de etapas en el formulario de nuevo proceso     
def comboEtapa(ti):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                                SELECT 
                                 *
                                FROM tbl_etapas
                                WHERE id_tipo_proceso = %s 
                            """)
                mycursor.execute(querySQL,(ti,))
                etapas = mycursor.fetchall()
                return etapas
    except Exception as e:
        print(f"Ocurrió un error al hacer la consulta de las etapas: {e}")
        return []

def datoCli(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                            SELECT 
                                e.id,
                                e.nombre_empleado,
                                e.apellido_empleado,
                                p.id_cliente,
                                p.id_proceso,
                                p.proceso_numero,
                                p.id_tipo_proceso
                            FROM tbl_empleados AS e, tbl_procesos AS p 
                          	WHERE e.id = p.id_cliente AND p.id_proceso =  %s
                            """)
                mycursor.execute(querySQL, (id,))
                etapas = mycursor.fetchall()
                return etapas
    except Exception as e:
        print(f"Ocurrió un error al hacer la consulta de las etapas: {e}")
        return []

# cargando combo de delitos en el formulario de nuevo proceso     
def comboDelitos():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                                SELECT 
                                *
                                FROM tbl_delitos
                                WHERE delito_estado = 1 
                            """)
                mycursor.execute(querySQL)
                delitos = mycursor.fetchall()
                return delitos
    except Exception as e:
        print(f"Ocurrió un error al hacer la consultar los delitos: {e}")
        return []
    
#cargando la lista de etapas de un proceso
def cargaListaEtapa(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                            SELECT 
                                p.id_proceso,
                                p.proceso_numero,
                                e.id_etapa,
                                e.nombre_etapa,
                                n.anotacion_etapa,
                                n.id_etapa_etapa,
                                n.id_new_etapa,
                                n.id_proceso_etapa,
                                n.fecha_audiencia,
                                n.estado_new_etapa
                            FROM
                                tbl_procesos AS P, tbl_etapas AS e, tbl_new_etapa as n
                            WHERE
                                n.estado_new_etapa=1 AND e.id_etapa=n.id_etapa_etapa AND n.id_proceso_etapa=p.id_proceso AND p.id_proceso= %s 
                            """)
                mycursor.execute(querySQL, (id,))
                etapaslist = mycursor.fetchall()
                return etapaslist
    except Exception as e:
        print(f"Ocurrió un error al hacer la consulta de las etapas: {e}")
        return []

#cargando la lista de documentos de un proceso
def cargaDocument(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                            SELECT 
                               *
                            FROM
                                tbl_carga_document AS c, tbl_procesos AS p, tbl_etapas AS e, tbl_empleados AS l
                            WHERE
                               c.id_proceso=p.id_proceso AND l.id=c.id_cliente AND e.id_etapa=c.id_etapa AND p.id_proceso=%s
                            """)
                mycursor.execute(querySQL, (id,))
                cdoc = mycursor.fetchall()
                return cdoc
    except Exception as e:
        print(f"Ocurrió un error al hacer la consulta de los documentos de un proceso: {e}")
        return []

# actualizar la informacion de un cliente
def procesar_actualizacion_form(dataForm, foto_perfil):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = dataForm['nombre_empleado']
                apellido_empleado = dataForm['apellido_empleado']
                sexo_empleado = dataForm['sexo_empleado']
                telefono_empleado = dataForm['telefono_empleado']
                ubicacion = dataForm['ubicacion']
                email_empleado = dataForm['email_empleado']
                id_empleado = dataForm['id_empleado']
                cod=dataForm['cod']
                if cod:
                    fotoForm = procesar_imagen_perfil(foto_perfil)
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            ubicacion = %s,
                            email_empleado = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, ubicacion, email_empleado, cod, id_empleado)
                else:
                    fotoForm = procesar_imagen_perfil(foto_perfil)
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            ubicacion = %s,
                            email_empleado = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, ubicacion, fotoForm, email_empleado, id_empleado)
                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()
        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None

# actualizar la informacion de un proceso
def procesar_actualizacion_proceso(dataForm):
    id=dataForm['id']
    idti=dataForm['proce']
    id_area_pro=dataForm['area']
    proceso_numero=dataForm['proceso_numero']
    demandado=dataForm['demandado']
    depto=dataForm['depto']
    id_juzgado=dataForm['id_juzgado']
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                        UPDATE tbl_procesos
                        SET 
                            id_tipo_proceso = %s,
                            id_area_proceso = %s,
                            proceso_numero = %s,
                            demandado = %s,
                            id_depto = %s,
                            id_juzgado = %s
                        WHERE id_proceso = %s
                    """
                values = (idti, id_area_pro, proceso_numero, demandado, depto, id_juzgado, id)
                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()
        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar la actualizacion de un proceso: {e}")
        return None

# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error procesar_actualizacion_formen lista_usuariosBD : {e}")
        return []

# Eliminar documento de identidad de un cliente al hacer una actualizacion
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)
                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []

# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
# inactivar etapa
def borrarEtapa(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "UPDATE tbl_new_etapa SET estado_new_etapa = 0 WHERE id_new_etapa = %s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
# inactivar proceso
def borrarProceso(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "UPDATE tbl_procesos SET estado = 0 WHERE id_proceso = %s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminar proceso : {e}")
        return []

# inactivar delito que se encuentra el el formulario de actualizar o editar
def borrarDelito(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "UPDATE tbl_delitos_agravantes SET estado = 0 WHERE id_delagra = %s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminar proceso : {e}")
        return []

# consultas de tipo proceso y area para llenar combobox depediente
def comboTipoProceso():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                                SELECT 
                                *
                                FROM tbl_tipo_proceso
                                WHERE proceso_estado = 1 
                            """)
                mycursor.execute(querySQL)
                combotproces = mycursor.fetchall()
                return combotproces
    except Exception as e:
        print(f"Ocurrió un error al hacer la consultar los tipo de proceso para el combo dependiente primer combo==>>: {e}")
        return []

# consultas de municipio para llenar los juzgados
def comboDepto():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                                SELECT 
                                *
                                FROM tbl_municipios
                            """)
                mycursor.execute(querySQL)
                combodepto = mycursor.fetchall()
                return combodepto
    except Exception as e:
        print(f"Ocurrió un error al hacer la consultar de los municipios: {e}")
        return []

#buscar agravantes que dependan de delitos
def buscarAgra(id_delito_agravante):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                                SELECT 
                                    *
                                FROM
                                     tbl_agravantes
                                WHERE
                                    id_delito_agravante=%s
                            """)
                mycursor.execute(querySQL, (id_delito_agravante,))
                agrava = mycursor.fetchall()
                return agrava
    except Exception as e:
        print(f"Ocurrió un error al hacer la consultar de los agravantes: {e}")
        return []
